import openpyxl
from openpyxl import Workbook

def open_excel_sheet(filename, sheet_name):
    wb = openpyxl.load_workbook(filename)
    sheet = wb.get_sheet_by_name(sheet_name)
    return sheet

def find_word_in_row(sheet, column_name, word):
    for row in range(1, sheet.max_row +1):
        if(sheet[column_name + str(row)].value == word):
            print('Inicio de la lista detectado en la fila {} del archivo'.format(row))
            list_detected=row
            break
    return list_detected

def find_word_in_column(sheet, row_index, word):
    for col in sheet.iter_cols(min_row=row_index, max_row =row_index):
        for cell in col:
            if(cell.value == word):
                print('Inicio de la lista detectado en la columna {} del archivo'.format(cell.coordinate))
                list_detected=cell.coordinate
                break
    return list_detected

def extract_list_by_index(sheet, row_start, column):
    students=[]
    for row in range (row_start+1, sheet.max_row+1):
        if(sheet[column + str(row)].value!= None):
            students.append(sheet[column + str(row)].value)
    students = list(dict.fromkeys(students)) #removing duplicates
    return students

def attended_split(list_students, attendance_students):
    #Use the oficial list to split attendance list in two list:
    #attended in list (attended_students) and attended out of list (attended_not_in_list)

    attended_students = []
    attended_not_in_list = attendance_students

    for student in list_students:
        #Split the name in official list
        words_name = student.lower().split(' ')
        word_find_count = 0
        for student_attended in attendance_students:
            #Iterate over attendance list searching the splited name
            for word in words_name:
                #Search in the current attendance item for words in the name splited
                if(student_attended.lower().find(word) != -1):
                    word_find_count += 1
                    if(word_find_count == 2):
                        #Store attended official students if at least 2 words are founded
                        # replace 2 with len(words_name) to verify all the words in the official list
                        # at the end of the process attended_not_in_list will contain the remained names 
                        #for people that attend the class but aren't in the official list.
                        attended_students.append(student)
                        attended_not_in_list.remove(student_attended)
                        break
    return attended_students, attended_not_in_list

def extract_absence_students(list_students, attended_students):
    absence_students = list_students[:]

    for item in attended_students:
        absence_students.remove(item)
    return absence_students

def write_column_report(sheet, column, title, data):
    sheet[column+'1']=title
    for row in range (0, len(data)):
        sheet[column + str(row+2)] = data[row]

def write_report(class_name, list_students, attended_students, attended_not_in_list, absence_students):
    
    columns_and_titles = []
    columns_and_titles.append(('A', 'Lista de estudiantes', list_students))
    columns_and_titles.append(('B', 'Asistentes en lista', attended_students))
    columns_and_titles.append(('C', 'Asistentes NO en lista', attended_not_in_list))
    columns_and_titles.append(('D', 'Ausentes', absence_students))

    try:
        wb = openpyxl.load_workbook("report.xlsx")
        sheet = wb.create_sheet("sheet_new")
        sheet.title = class_name
    except:
        print('No se encuentra report.xlsx, se creara este archivo')
        wb = Workbook()
        sheet = wb.active
        sheet.title = class_name

    for item in columns_and_titles:
        write_column_report(sheet, item[0], item[1], item[2])
    wb.save("report.xlsx")

if __name__ == "__main__":
    class_list=['matematicas'] #Agregar los cursos necesarios

    list_sheet_name='Sheet1'
    list_column='B'
    list_word='Apellidos y Nombres'

    attendance_file='Meet.xlsx'
    attendance_sheet_name='Attendance'

    for class_name in class_list:
        print('########## Iniciando verificación de asistencia de {}'.format(class_name))
        print('########## Iniciando detección de lista oficial de {}'.format(class_name))
        list_sheet = open_excel_sheet(class_name+'.xlsx',list_sheet_name)
        list_row_index = find_word_in_row(list_sheet, list_column, list_word)
        list_students = extract_list_by_index(list_sheet, list_row_index, list_column)

        print('########## Iniciando busqueda de columna de asistencia de {}'.format(class_name))
        attendance_sheet=open_excel_sheet(attendance_file,attendance_sheet_name)
        attendance_column_index = find_word_in_column(attendance_sheet,1,class_name)
        attendance_students = extract_list_by_index(attendance_sheet, 2, attendance_column_index[0:-1])

        print('########## Iniciando comparación de lista {}'.format(class_name))
        attended_students, attended_not_in_list = attended_split(list_students, attendance_students)
        absence_students = extract_absence_students(list_students, attended_students)

        print('########## Escribiendo el reporte en report.xlsx en la hoja {}'.format(class_name))
        write_report(class_name, list_students, attended_students, attended_not_in_list, absence_students)